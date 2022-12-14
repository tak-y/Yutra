import json

from fastapi import FastAPI
import datetime

import sqlite3
import os
import sys
import json
import openpyxl
import jpholiday
import googlemaps
import pandas as pd
import requests
import math
from geopy.distance import geodesic
import time
import geocoder
from transport_nearest_time import transport_nearest_origtime,transport_nearest_desttime
app = FastAPI()

@app.get("/")
def read():
    return {"Hello": "peko"}

@app.get("/time_depart/{depart_stop}/{depart_pos}/{arrival_stop}/{arrival_pos}/{depart_year}/{depart_month}/{depart_day}/{depart_hour}/{depart_minutes}/{prepare_minutes}/{arrival_year}/{arrival_month}/{arrival_day}/{arrival_hour}/{arrival_minutes}}/{arrival_allow_minutes}")
def lookup(depart_stop:str,depart_pos:str,arrival_stop:str,arrival_pos:str,depart_year:int,depart_month:int,depart_day:int,depart_hour:int,depart_minutes:int,prepare_minutes:int,arrival_year:int,arrival_month:int,arrival_day:int,arrival_hour:int,arrival_minutes:int,arrival_allow_minutes:int,event, context):
    import datetime
    global count
    global o
    global d
    global o_O_lat
    global o_O_lng
    global d_O_lat
    global d_O_lng
    global selected_agency
    global select_year
    global select_month
    global select_day
    global select_hour
    global select_minutes
    global select_margin_minutes
    global select_DATE
    googleapikey = 'AIzaSyC8rTI8Yv1LrEDnRJ109DfpUsdaQBXAfhE'
    gmaps = googlemaps.Client(key=googleapikey)
####

    def geo_google():
        url = 'https://www.googleapis.com/geolocation/v1/geolocate?key=AIzaSyC8rTI8Yv1LrEDnRJ109DfpUsdaQBXAfhE'
        data = {"considerIp": "true",
                }
        headers = {
            'Content-Type': 'application/json',
        }
        req = requests.post(url, json.dumps(data).encode(), headers)
        body = req.json()
        return body
    #if current_dep==True
    def closest_o(data,o):
        return min(data, key=lambda p: geodesic((o['?ܓx'], o['?o?x']), (p['?ܓx'], p['?o?x'])).m, default="")
    def closest_d(data,d):
        return min(data, key=lambda p: geodesic((d['?ܓx'], d['?o?x']), (p['?ܓx'], p['?o?x'])).m, default="")
    #if current_dep==True
    def nearest_stop_orig_lat(data,o1):
        return min(data,key=lambda p:geodesic((o1['?ܓx'], o1['?o?x']), (p['?ܓx'], p['?o?x'])).m, default="")
            
    def nearest_stop_dest_lat(data,d1):
        return min(data,key=lambda p:geodesic((d1['?ܓx'], d1['?o?x']), (p['?ܓx'], p['?o?x'])).m, default="")
            
    def sheet_delete_row():
        for col in ws_select_bus_agency.iter_cols(min_row=4,max_row=ws_select_bus_agency.max_row,min_col=1,max_col=ws_select_bus_agency.max_column):
            for cell in col:
                if ws_select_bus_agency.cell(row=cell.row,column=2).value==0:
                    ws_select_bus_agency.delete_rows(cell.row)
                    print(cell.row)
                    return ws_select_bus_agency
                else:
                    return ws_select_bus_agency

    def select_o_d(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide):
        global selected_agency
        global select_margin_minutes
        global transport_orig_datetime
        global transport_orig_date
        global orig_time
        global dest_time
        global margin_time
        tram_stop=pd.read_csv('tram_stop.csv',encoding='cp932',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
        tram_stop=tram_stop[(tram_stop['?ܓx']!='-')|(tram_stop['?o?x']!='-')]
        tram_stop=tram_stop.astype({'?ܓx':float,'?o?x':float})
        tram_stop=tram_stop.query('-90.0<?ܓx<90.0|?o?x!=0')
        tram_stop=tram_stop.to_dict(orient='records')
        api_o_tram='https://api.mapbox.com/directions/v5/mapbox/{transport}/{o_lng},{o_lat};{lng},{lat}?geometries=geojson&access_token=pk.eyJ1IjoidGFrLXkiLCJhIjoiY2tnbjFpN3RiMDMwczM3bXNkem9sbm5zZCJ9.TK7AsKUUkR0kicGCyFWBsQ'
        transport='walking'
        print(o)
        url_o_tram=api_o_tram.format(transport=transport,o_lat=o_O_lat,o_lng=o_O_lng,lat=o['?ܓx'],lng=o['?o?x']) if len(depart_stop)>0 or len(depart_pos)>0 else api_o_tram.format(transport=transport,o_lat=geo_google()["location"]["lat"],o_lng=geo_google()["location"]["lng"],lat=o['?ܓx'],lng=o['?o?x'])
        response_o_tram = requests.get(url_o_tram)
        data_o_tram = response_o_tram.json()
        route_o_tram = data_o_tram['routes'][0]['geometry']['coordinates']
        lng_o_tram = []
        lat_o_tram = []
        for i in range(len(route_o_tram)-1):
            lng_o_tram.append(route_o_tram[i][1])
            lat_o_tram.append(route_o_tram[i][0])
        mylist_o_tram = []
        for i in range(len(route_o_tram)-1):
            mylist_o_tram.append([lat_o_tram[i], lng_o_tram[i]])
        orig_time = math.ceil(data_o_tram['routes'][0]['duration'] / 60)
        print("?o?????????܂ł̎??Ԃ?")
        print(orig_time)
        api_d_tram='https://api.mapbox.com/directions/v5/mapbox/{transport}/{d_lng},{d_lat};{lng},{lat}?geometries=geojson&access_token=pk.eyJ1IjoidGFrLXkiLCJhIjoiY2tnbjFpN3RiMDMwczM3bXNkem9sbm5zZCJ9.TK7AsKUUkR0kicGCyFWBsQ'

        url_d_tram=api_d_tram.format(transport=transport,d_lat=d_O_lat,d_lng=d_O_lng,lat=d['?ܓx'],lng=d['?o?x'])
        response_d_tram = requests.get(url_d_tram)
        data_d_tram = response_d_tram.json()
        print(url_d_tram)
        print(data_d_tram)
        route_d_tram = data_d_tram['routes'][0]['geometry']['coordinates']
        lng_d_tram = []
        lat_d_tram = []
        for i in range(len(route_d_tram)-1):
            lng_d_tram.append(route_d_tram[i][1])
            lat_d_tram.append(route_d_tram[i][0])
        mylist_d_tram = []
        for i in range(len(route_d_tram)-1):
            mylist_d_tram.append([lat_d_tram[i], lng_d_tram[i]])
        dest_time = math.ceil(data_d_tram['routes'][0]['duration'] / 60)
        if len(depart_year)>0:
            select_depart=o['?◯??']
            select_year=int(depart_year)
            select_month=int(depart_month)
            select_day=int(depart_day)
            select_hour=int(depart_hour)
            select_minutes=int(depart_minutes)
            select_margin_minutes=int(prepare_minutes)
            select_date=datetime.datetime.strptime(str(select_year)+'-'+str(select_month)+'-'+str(select_day),"%Y-%m-%d")
            select_time=datetime.datetime.strptime(str(select_hour)+':'+str(select_minutes),"%H:%M")+datetime.timedelta(minutes=-select_margin_minutes)
                
            select_DATE=str(select_year)+str(select_month)+str(select_day)
            DATE = int(select_DATE) # ???t?͂W?????????̌`??
            Date = datetime.date(int(select_year), int(select_month), int(select_day))
            if int(select_hour)>1:
                transport_orig_date=datetime.datetime(year=select_year, month=select_month, day=select_day, hour=int(select_hour),minute=int(select_minutes))
                print("?ݒ肵?????Ԃ?")
                print(transport_orig_date)
                print("?A?N?Z?X???ԁ{????????")
                print(select_margin_minutes+int(orig_time))
                if depart_pos != "?o???{??":
                    transport_orig_datetime=transport_orig_date-datetime.timedelta(minutes = int(orig_time))
                    print('?O?o???????????Ԃ?{}'.format(transport_orig_datetime))
                    margin_time=select_margin_minutes+int(orig_time)
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif depart_stop != "?o????????":
                    transport_orig_datetime=transport_orig_date
                    margin_time=select_margin_minutes
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif arrival_stop != "?ړI??????":
                    transport_orig_datetime=transport_orig_date
                    margin_time=select_margin_minutes
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif arrival_pos != "?ړI?{??":
                    transport_orig_datetime=transport_orig_date-datetime.timedelta(int(dest_time))
                    margin_time=select_margin_minutes+int(dest_time)
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                else:
                    print(None)

            else:
                pass
                    
        elif len(arrival_year)>0:
            select_depart=o['?◯??']
            select_year=int(arrival_year)
            select_month=int(arrival_month)
            select_day=int(arrival_day)
            select_hour=int(arrival_hour)
            select_minutes=int(arrival_minutes)
            select_margin_minutes=int(arrival_allow_minutes)
            select_date=datetime.datetime.strptime(str(select_year)+'-'+str(select_month)+'-'+str(select_day),"%Y-%m-%d")
            select_time=datetime.datetime.strptime(str(select_hour)+':'+str(select_minutes),"%H:%M")+datetime.timedelta(minutes=-select_margin_minutes)
            select_DATE=str(select_year)+str(select_month)+str(select_day)
            DATE = int(select_DATE) # ???t?͂W?????????̌`??
            Date = datetime.date(int(select_year), int(select_month), int(select_day))
            transport_orig=select_time
            select_dest=d['?◯??']
            if int(select_hour)>1:
                transport_orig_date=datetime.datetime(year=select_year, month=select_month, day=select_day, hour=int(select_hour),minute=int(select_minutes))
                if depart_pos != "?o???{??":
                    transport_orig_datetime=transport_orig_date-datetime.timedelta(minutes=int(orig_time))
                    print('?O?o???????????Ԃ?{}'.format(transport_orig_datetime))
                    margin_time=select_margin_minutes+int(orig_time)
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif depart_stop != "?o????????":
                    transport_orig_datetime=transport_orig_date
                    margin_time=select_margin_minutes
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif arrival_stop != "?ړI??????":
                    transport_orig_datetime=transport_orig_date
                    margin_time=select_margin_minutes
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                elif arrival_pos != "?ړI?{??":
                    transport_orig_datetime=transport_orig_date-datetime.timedelta(minutes=int(dest_time))
                    margin_time=select_margin_minutes+int(dest_time)
                    excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime)
                else:
                    print(None)

            else:
               pass
        else:
            pass
        

    def excel_select(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide,margin_time,transport_orig_datetime):
        global selected_agency
        global wheel
        import datetime
        global before_time
        global after_time
        global prepare_want_time
        global ride_time
        global select_margin_minutes
        before_time=True if before_time=='1?֑O?ɂ???' else False
        after_time=True if after_time=='1?֌??ɂ???' else False
        tram_stop=pd.read_csv('tram_stop.csv',encoding='cp932',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
        tram_stop=tram_stop[(tram_stop['?ܓx']!='-')|(tram_stop['?o?x']!='-')]
        tram_stop=tram_stop.astype({'?ܓx':float,'?o?x':float})
        tram_stop=tram_stop.query('-90.0<?ܓx<90.0|?o?x!=0')
        tram_stop=tram_stop.to_dict(orient='records')
        def transport_nearest_origtime(data):
            print(before_time)
            print(after_time)
            global transport_orig_datetime
            global orig_time
            global margin_time
            print(orig_time)
            print(select_margin_minutes)
            transport_orig_datetime0=transport_orig_datetime
            
            print('?o???????????Ԃ?{}'.format(transport_orig_datetime0))
            for i in range(0,len(data)-1):
                if transport_orig_datetime0 > datetime.datetime(data[i][0],data[i][1],data[i][2],data[i][3],data[i][4]):
                    min_detetime=datetime.datetime(data[i][0],data[i][1],data[i][2],data[i][3],data[i][4])
                else:
                    pass
            return min_detetime

                                                
        def transport_nearest_desttime(data,date):
            print(before_time)
            print(after_time)
            global transport_orig_datetime
            global orig_time
            global margin_time
            print(orig_time)
            print('?o???????????Ԃ?{}'.format(transport_orig_datetime))
            print(date)
            dest_idx=date
            min_detetime=datetime.datetime(data[dest_idx][0],data[dest_idx][1],data[dest_idx][2],data[dest_idx][3],data[dest_idx][4])

            return min_detetime
        def transport_nearest_arrival_prefer_origtime(data,date):
            print(before_time)
            print(after_time)
            global transport_orig_datetime
            global orig_time
            global margin_time
            print(orig_time)
            print(select_margin_minutes)
            print('?o?????Ȃ??Ƃ????Ȃ????Ԃ?')
            print(date)
            delta=[]
            dest_idx=date
            min_detetime=datetime.datetime(data[dest_idx][0],data[dest_idx][1],data[dest_idx][2],data[dest_idx][3],data[dest_idx][4])
            return min_detetime

                                                
        def transport_nearest_arrival_prefer_desttime(data):
            print(before_time)
            print(after_time)
            global transport_orig_datetime
            global orig_time
            global margin_time
            print(orig_time)
            transport_orig_datetime0=transport_orig_datetime
            print('?????????????Ԃ?{}'.format(transport_orig_datetime0))
            for i in range(0,len(data)-1):
                if transport_orig_datetime0 > datetime.datetime(data[i][0],data[i][1],data[i][2],data[i][3],data[i][4]):
                    min_detetime=datetime.datetime(data[i][0],data[i][1],data[i][2],data[i][3],data[i][4])
                    print(min_detetime)
                else:
                    pass
            return min_detetime
            
        try:
            select_depart=o['?◯??']
            select_year=int(depart_year)
            select_month=int(depart_month)
            select_day=int(depart_day)
            select_hour=int(depart_hour)
            select_minutes=int(depart_minutes)
            select_margin_minutes=int(prepare_minutes)
            select_date=datetime.datetime.strptime(str(select_year)+'-'+str(select_month)+'-'+str(select_day),"%Y-%m-%d")
            select_time=datetime.datetime.strptime(str(select_hour)+':'+str(select_minutes),"%H:%M")+datetime.timedelta(minutes=-select_margin_minutes)
                
            select_DATE=str(select_year)+str(select_month)+str(select_day)
            DATE = int(select_DATE) # ???t?͂W?????????̌`??
            print(DATE)
            Date = datetime.date(int(select_year), int(select_month), int(select_day))
        except:
            select_depart=o['?◯??']
            select_year=int(arrival_year)
            select_month=int(arrival_month)
            select_day=int(arrival_day)
            select_hour=int(arrival_hour)
            select_minutes=int(arrival_minutes)
            select_margin_minutes=int(arrival_allow_minutes)
            select_date=datetime.datetime.strptime(str(select_year)+'-'+str(select_month)+'-'+str(select_day),"%Y-%m-%d")
            select_time=datetime.datetime.strptime(str(select_hour)+':'+str(select_minutes),"%H:%M")+datetime.timedelta(minutes=-select_margin_minutes)
            select_DATE=str(select_year)+str(select_month)+str(select_day)
            DATE = int(select_DATE) # ???t?͂W?????????̌`??
            print(DATE)
            Date = datetime.date(int(select_year), int(select_month), int(select_day))
      
        bus_agency={'???َs???ʋ?':['Hakodate_tram_weekday.xlsx','Hakodate_tram_holiday.xlsx'],'???ˎs?R?~???j?e?B?o?X':['seto_bus.csv']}

        wb_select_bus_agency=openpyxl.load_workbook('Hakodate_tram_holiday.xlsx') if selected_agency=='???َs???ʋ?' else None
        ws_select_bus_agency_1 = wb_select_bus_agency['Table 1']
        ws_select_bus_agency_2 = wb_select_bus_agency['Table 2']
        print('?ݒ肵????????')
        print(transport_orig_datetime)
        print(o)
        print(d)
        wheel = True if wheel=='?Ԉ֎q?g?p' else False
        candidate_orig_time=[]
        candidate_dest_time=[]
        transport_orig_timetable=[]
        transport_dest_timetable=[]
        print('orig')
        print(d['?ܓx'])
        o_cell_row=o_cell_row_decide
        d_cell_row=d_cell_row_decide
        if ws_select_bus_agency.cell(row=o_cell_row,column=2).value == o['?ܓx']:
            print("____")
            wb_transport=openpyxl.load_workbook(bus_agency['???َs???ʋ?'][1])
            ws0_transport=wb_transport['Table 1'] if ws_select_bus_agency.title=="Table 1" else wb_transport['Table 2']
            if Date.weekday() >= 5 or jpholiday.is_holiday(Date):
                o_cell_row=o_cell_row_decide
                d_cell_row=d_cell_row_decide
                url="{agency}"
                transport_stop=pd.read_excel(url.format(agency=bus_agency[selected_agency][1]), index_col=0, sheet_name=0)

                print('holiday')
                print("____")
                if wheel ==True:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour= str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')]) if len(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)>0 else 0
                                minute= str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:]) if len(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value) > 0 else 0
                                            
                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour= str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')]) if len(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value) > 0 else 0
                                minute= str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:]) if len(ws0_transport.cell(row=int(d_cell_row),column=cell.column).valuee) > 0 else 0
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                        
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                                
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time:
                            if len(transport_orig_timetable[k])>0:
                                candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                                candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                            else:
                                candidate_orig_time.append([0,0,0,0,0])
                                candidate_dest_time.append([0,0,0,0,0])
                            
                        else:
                            pass
                                    
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        return {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                    else:
                                    
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print("____")
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        Sm.current = 'Timer_set'
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        return {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                else:
                    print("____")
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])
                                
                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time:
                            if len(transport_orig_timetable[k])>0:
                                candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                                candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                            else:
                                candidate_orig_time.append([0,0,0,0,0])
                                candidate_dest_time.append([0,0,0,0,0])
                        else:
                            pass
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time+orig_time)
                        depart_want_time=ride_time-datetime.timedelta(minutes=orig_time)
                        remnant_hour=abs(prepare_want_time-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(prepare_want_time-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        print('You should prepare at {}'.format(depart_want_time))
                        print('margintime is {}'.format(margin_time))
                        print('You should ride on {}'.format(ride_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        return {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                    else:
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        depart_want_time=ride_time
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        return {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
            else:
                url="{agency}"
                transport_stop=pd.read_excel(url.format(agency=bus_agency[selected_agency][0]), index_col=0, sheet_name=0)
                wb_transport=openpyxl.load_workbook(bus_agency['???َs???ʋ?'][0])
                o_cell_row=o_cell_row_decide
                d_cell_row=d_cell_row_decide
                print('weekday')

                if wheel ==True:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])

                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                              
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for k in range(0,len(transport_orig_timetable)-1):
                        print(transport_orig_timetable[k])
                        if transport_orig_timetable[k] not in candidate_orig_time and transport_dest_timetable[k] not in candidate_dest_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                            
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    if len(depart_year) > 0:
                        print(candidate_orig_time)
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        print("____")
                        print(candidate_orig_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time+int(prepare_minutes))
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        remnant_hour=abs(prepare_want_time-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(prepare_want_time-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        print('You should prepare at {}'.format(depart_want_time))
                        print('margintime is {}'.format(margin_time))
                        print('You should ride on {}'.format(ride_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                    else:
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time+int(arrival_allow_minutes))
                        depart_want_time=ride_time
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }

                else:
                    print(o_cell_row)
                    print(d_cell_row)
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])

                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                              
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for k in range(0,len(transport_orig_timetable)-1):
                        print(transport_orig_timetable[k])
                        if transport_orig_timetable[k] not in candidate_orig_time and transport_dest_timetable[k] not in candidate_dest_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                            
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time+orig_time)
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        remnant_hour=abs(prepare_want_time-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(prepare_want_time-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        print('You should prepare at {}'.format(depart_want_time))
                        print('margintime is {}'.format(margin_time))
                        print('You should ride on {}'.format(ride_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }

                    else:
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        prepare_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
        else:
            d_temp_cell_row=cell_o
            print('orig')
            print(d['?ܓx'])
            ws_select_bus=[]
            for i in range(2,27):
                ws_select_bus.append(i)
            o_cell_row=[i for i in ws_select_bus if ws_select_bus_agency_2.cell(row=i,column=2).value== o['?ܓx']]
            print(ws_select_bus_agency_2.cell(row=o_cell_row,column=2).value)
            print('orig1')
            ws_select_bus=[]
            for j in range(2,27):
                ws_select_bus.append(j)
            d_cell_row=[j for j in ws_select_bus if ws_select_bus_agency_2.cell(row=j,column=2).value== d['?ܓx']]
            print(d_cell_row[0])
            print('dest1')
            if Date.weekday() >= 5 or jpholiday.is_holiday(Date):
                url="{agency}"
                transport_stop=pd.read_excel(url.format(agency=bus_agency[selected_agency][1]), index_col=0, sheet_name=0)
                wb_transport=openpyxl.load_workbook(bus_agency['???َs???ʋ?'][1])
                ws0_transport=wb_transport[wb_transport.sheetnames[0]]
                print('holiday')
                if wheel ==True:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])
                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)

                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time and transport_dest_timetable[k] not in candidate_dest_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                                    
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                    else:
                                    
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        print(candidate_orig_time)
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                else:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])

                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                                        
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                                                        
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    print('transport')
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                    else:
                                
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        print(candidate_orig_time)
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
            else:
                url="{agency}"
                transport_stop=pd.read_excel(url.format(agency=bus_agency[selected_agency][0]), index_col=0, sheet_name=0)
                print(bus_agency['???َs???ʋ?'][0])
                wb_transport=openpyxl.load_workbook(bus_agency['???َs???ʋ?'][0])
                ws0_transport=wb_transport['Table 1']
                print('weekday')
                if wheel ==True:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])

                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if cell.fill.fgColor.rgb == 'FFFF0000' and ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                            
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                                    
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                    else:
                                    
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        print(candidate_orig_time)
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        print("index")
                        print(candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                else:
                    for col in ws0_transport.iter_cols(min_col=4,min_row=o_cell_row,max_row=o_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_orig_timetable.count(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value).find(':')+1:])

                                transport_orig_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+str(hour)+','+str(minute))
                                transport_orig_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_orig_dict={transport_orig_route:str(ws0_transport.cell(row=int(o_cell_row),column=cell.column).value)}
                            else:
                                pass
                    for col in ws0_transport.iter_cols(min_col=4,min_row=d_cell_row,max_row=d_cell_row):
                        for cell in col:
                            if ws0_transport.cell(row=int(o_cell_row),column=cell.column).value != None and ws0_transport.cell(row=int(d_cell_row),column=cell.column).value != None and transport_dest_timetable.count(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)==0:
                                hour=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')])
                                minute=str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value[str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value).find(':')+1:])
                                transport_dest_timetable.append(str(select_year)+','+str(select_month)+','+str(select_day)+','+hour+','+minute)
                                            
                                transport_dest_route=str(ws0_transport.cell(row=34,column=cell.column).value)
                                transport_dest_dict={transport_dest_route:str(ws0_transport.cell(row=int(d_cell_row),column=cell.column).value)}
                            else:
                                pass
                    print(len(transport_orig_timetable))
                    print(len(transport_dest_timetable))
                    for k in range(0,len(transport_orig_timetable)-1):
                        if transport_orig_timetable[k] not in candidate_orig_time:
                            candidate_orig_time.append([int(transport_orig_timetable[k].split(',')[0]),int(transport_orig_timetable[k].split(',')[1]),int(transport_orig_timetable[k].split(',')[2]),int(transport_orig_timetable[k].split(',')[3]),int(transport_orig_timetable[k].split(',')[4])])
                            candidate_dest_time.append([int(transport_dest_timetable[k].split(',')[0]),int(transport_dest_timetable[k].split(',')[1]),int(transport_dest_timetable[k].split(',')[2]),int(transport_dest_timetable[k].split(',')[3]),int(transport_dest_timetable[k].split(',')[4])])
                                    
                        else:
                            candidate_orig_time.append([0,0,0,0,0])
                            candidate_dest_time.append([0,0,0,0,0])
                    if len(depart_year) > 0:
                        orig_stop_time=transport_nearest_origtime(candidate_orig_time)
                        print(orig_stop_time)
                        dest_stop_orig_prefer_time_l=[orig_stop_time.year,orig_stop_time.month,orig_stop_time.day,orig_stop_time.hour,orig_stop_time.minute]
                        dest_stop_time=transport_nearest_desttime(candidate_dest_time,candidate_orig_time.index(dest_stop_orig_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }
                    else:
                                    
                        dest_stop_arrival_prefer_time=transport_nearest_arrival_prefer_desttime(candidate_dest_time)
                        print(dest_stop_arrival_prefer_time)
                        print("____")
                        dest_stop_arrival_prefer_time_l=[dest_stop_arrival_prefer_time.year,dest_stop_arrival_prefer_time.month,dest_stop_arrival_prefer_time.day,dest_stop_arrival_prefer_time.hour,dest_stop_arrival_prefer_time.minute]
                        orig_stop_time=transport_nearest_arrival_prefer_origtime(candidate_orig_time,candidate_dest_time.index(dest_stop_arrival_prefer_time_l))
                        remnant_hour=abs(transport_orig_datetime-datetime.datetime.now()).seconds//(60*60)
                        remnant_minute=(abs(transport_orig_datetime-datetime.datetime.now()).seconds-remnant_hour*(60*60))/60
                        ride_time=orig_stop_time
                        depart_want_time=ride_time-datetime.timedelta(minutes=margin_time)
                        print('You should be about to prepare{}'.format(orig_stop_time))
                        print('You will arrive at{}'.format(dest_stop_arrival_prefer_time))
                        print(abs(transport_orig_datetime-datetime.datetime.now()).seconds/60)
                        
                        arrival_train = str(dest_stop_arrival_prefer_time)+'?ɓ????\????'
                        ride_train = ride_time.strftime('%Y?N%m??%d??%H??%M??')+'???ɏ??邽?߂?'
                        prepare_time = str(prepare_want_time)+'???珀??????'
                        depart_time = str(depart_want_time)+'?܂łɏo?????܂??B'
                        rest_time = '???????????͗]?T?̂????o???܂Ŏc??'+str(int(remnant_hour))+'????'+str(int(remnant_minute))
                        hour=str(remnant_hour)
                        minutes=str(math.ceil(remnant_minute))
                        seconds=str(00)
                        data = {"arrival_train":arrival_train,"ride_train":ride_train,"prepare_time":prepare_time,"depart_time":depart_time,"rest_time":rest_time,"hour":hour,"minutes":minutes,"seconds":seconds}
                        return {
                            'statusCode': 200,
                            'body': json.dumps(data)
                        }


##############??
    key="pk.eyJ1IjoidGFrLXkiLCJhIjoiY2tnbjFpN3RiMDMwczM3bXNkem9sbm5zZCJ9.TK7AsKUUkR0kicGCyFWBsQ"
    o_name = depart_stop if len(depart_stop)>0 else depart_pos if len(depart_pos)>0 else None
    d_name = arrival_stop if len(arrival_stop)>0 else arrival_pos if len(arrival_pos)>0 else None
    orig = geocoder.mapbox(o_name,key=key)
    dest = geocoder.mapbox(d_name,key=key)
    print(orig)
    o_O_lat = orig.lat
    o_O_lng = orig.lng
    d_O_lat = dest.lat
    d_O_lng = dest.lng
    print(d_O_lat)
    o1 = {'?{?ݖ?': o_name,'?ܓx':o_O_lat if len(depart_stop)>0 or len(depart_pos)>0 else geo_google()["location"]["lat"],'?o?x': o_O_lng if len(depart_stop)>0 or len(depart_pos)>0 else geo_google()["location"]["lng"]}
    print(o1)
    d1 = {'?{?ݖ?': d_name, '?ܓx': d_O_lat,'?o?x': d_O_lng}
    print(d1)
    tram_stop=pd.read_csv('tram_stop.csv',encoding='cp932',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
    tram_stop=tram_stop[(tram_stop['?ܓx']!='-')|(tram_stop['?o?x']!='-')]
    tram_stop=tram_stop.query('0<?ܓx<90.0|?o?x!=0|?o?x!=None')
    tram_stop=tram_stop.astype({'?ܓx':float,'?o?x':float})
    tram_stop=tram_stop.to_dict(orient='records')
    print(tram_stop[-1]['?ܓx'])
    
    selected_agency='???َs???ʋ?' if closest_o(tram_stop,o1)['???Ǝ?']=='???َs???ʋ?' and closest_d(tram_stop,d1)['???Ǝ?']=='???َs???ʋ?' else '???ˎs?R?~???j?e?B?o?X' if closest_o(tram_stop,o1)['???Ǝ?']=='???ˎs?R?~???j?e?B?o?X' and closest_d(tram_stop,d1)['???Ǝ?']=='???ˎs?R?~???j?e?B?o?X' else None
    tram_pd= pd.read_csv('Hakodate_tram_route.csv',encoding='UTF-8',header=None,names=['?o?X?▼','?ܓx','?o?x','???l']).fillna('-').replace(' ',0) if selected_agency=='???َs???ʋ?' else pd.read_csv('tram_stop.csv',encoding='UTF-8',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
    
    current_dep=True if o_name == None else False
    dep_stop=True if len(depart_stop)>0 else False
    dep_pos=True if len(depart_pos)>0 else False
    arr_stop=True if len(arrival_stop)>0 else False
    arr_pos=True if len(arrival_pos)>0 else False

    if selected_agency=='???َs???ʋ?':
        wb_select_bus_agency=openpyxl.load_workbook('Hakodate_tram_holiday.xlsx')
        ws_select_bus_agency0 = wb_select_bus_agency["Table1 route"]
        ws_select_bus_agency2 = wb_select_bus_agency["Table2 route"]
        ws_select_bus_agency0_time = wb_select_bus_agency["Table 1"]
        ws_select_bus_agency2_time = wb_select_bus_agency["Table 2"]
        tram_pd=tram_pd.to_dict(orient='records')
        near_st_orig_lat1=nearest_stop_orig_lat(tram_pd,o1)['?ܓx']
        near_st_orig_lng1=nearest_stop_orig_lat(tram_pd,o1)['?o?x']
        near_st_dest_lat1=nearest_stop_dest_lat(tram_pd,d1)['?ܓx']
        near_st_dest_lng1=nearest_stop_dest_lat(tram_pd,d1)['?o?x']
        print(near_st_orig_lat1)
        print(near_st_orig_lng1)
        print(near_st_dest_lat1)
        print(near_st_dest_lng1)
        ws_select_orig=near_st_orig_lat1
        ws_select_dest=near_st_dest_lat1
        tram_pd= pd.read_csv('Hakodate_tram_route.csv',encoding='UTF-8',header=None,names=['?o?X?▼','?ܓx','?o?x','???l']).fillna('-').replace(' ',0) if selected_agency=='???َs???ʋ?' else pd.read_csv('tram_stop.csv',encoding='UTF-8',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
    
        hakodate_lat_orig=tram_pd.query('?ܓx!=@ws_select_orig')
        hakodate_lat_dest=tram_pd.query('?ܓx!=@ws_select_dest')
        hakodate_lat_orig=hakodate_lat_orig.to_dict(orient='records')
        hakodate_lat_dest=hakodate_lat_dest.to_dict(orient='records')
        near_st_orig_lat2=nearest_stop_orig_lat(hakodate_lat_orig,o1)['?ܓx']
        near_st_orig_lng2=nearest_stop_orig_lat(hakodate_lat_orig,o1)['?o?x']
        near_st_dest_lat2=nearest_stop_dest_lat(hakodate_lat_dest,d1)['?ܓx']
        near_st_dest_lng2=nearest_stop_dest_lat(hakodate_lat_dest,d1)['?o?x']
        print(near_st_orig_lat2)
        print(near_st_orig_lng2)
        print(near_st_dest_lat2)
        print(near_st_dest_lng2)
        ws_select_bus=[]
        for j in range(2,28):
            ws_select_bus.append(j)
        o_cell_row1=[j for j in ws_select_bus if ws_select_bus_agency0.cell(row=j,column=2).value== near_st_orig_lat1]
        o_cell_row2=[j for j in ws_select_bus if ws_select_bus_agency0.cell(row=j,column=2).value== near_st_orig_lat2]
        d_cell_row1=[j for j in ws_select_bus if ws_select_bus_agency0.cell(row=j,column=2).value== near_st_dest_lat1]
        d_cell_row2=[j for j in ws_select_bus if ws_select_bus_agency0.cell(row=j,column=2).value== near_st_dest_lat2]
        o_cell_row3=[j for j in ws_select_bus if ws_select_bus_agency2.cell(row=j,column=2).value== near_st_orig_lat1]
        o_cell_row4=[j for j in ws_select_bus if ws_select_bus_agency2.cell(row=j,column=2).value== near_st_orig_lat2]
        d_cell_row3=[j for j in ws_select_bus if ws_select_bus_agency2.cell(row=j,column=2).value== near_st_dest_lat1]
        d_cell_row4=[j for j in ws_select_bus if ws_select_bus_agency2.cell(row=j,column=2).value== near_st_dest_lat2]
        print(o_cell_row1)
        print(o_cell_row2)
        print(d_cell_row1)
        print(d_cell_row2)
        print(o_cell_row3)
        print(o_cell_row4)
        print(d_cell_row3)
        print(d_cell_row4)
        try:
            if len(o_cell_row1)>0 and len(d_cell_row1)>0 and o_cell_row1[0]<d_cell_row1[0]:
                o_cell_row_decide=o_cell_row1[0]
                d_cell_row_decide=d_cell_row1[0]
                ws_select_bus_agency=ws_select_bus_agency0_time
                print("1")
                print(ws_select_bus_agency)
            else:
                if len(o_cell_row1)>0 and len(d_cell_row2)>0 and o_cell_row1[0]<d_cell_row2[0]:
                    o_cell_row_decide=o_cell_row1[0]
                    d_cell_row_decide=d_cell_row2[0]
                    ws_select_bus_agency=ws_select_bus_agency0_time
                    print("2")
                    print(ws_select_bus_agency)
                else:
                    if len(o_cell_row2)>0 and len(d_cell_row2)>0 and o_cell_row2[0]<d_cell_row2[0]:
                        o_cell_row_decide=o_cell_row2[0]
                        d_cell_row_decide=d_cell_row2[0]
                        ws_select_bus_agency=ws_select_bus_agency0_time
                        print("3")
                        print(ws_select_bus_agency)
                    else:
                        if len(o_cell_row3)>0 and len(d_cell_row3)>0 and o_cell_row3[0]<d_cell_row3[0]:
                            o_cell_row_decide=o_cell_row3[0]
                            d_cell_row_decide=d_cell_row3[0]
                            ws_select_bus_agency=ws_select_bus_agency2_time
                            print("4")
                            print(ws_select_bus_agency)
                        else:
                            if len(o_cell_row3)>0 and len(d_cell_row4)>0 and o_cell_row3[0]<d_cell_row4[0]:
                                o_cell_row_decide=o_cell_row3[0]
                                d_cell_row_decide=d_cell_row4[0]
                                ws_select_bus_agency=ws_select_bus_agency2_time
                                print("5")
                                print(ws_select_bus_agency)
                            else:
                                if len(o_cell_row4)>0 and len(d_cell_row3)>0 and o_cell_row4[0]<d_cell_row3[0]:
                                    o_cell_row_decide=o_cell_row4[0]
                                    d_cell_row_decide=d_cell_row3[0]
                                    ws_select_bus_agency=ws_select_bus_agency2_time
                                    print("6")
                                    print(ws_select_bus_agency)
                                else:
                                    if len(o_cell_row4)>0 and len(d_cell_row4)>0 and o_cell_row4[0]<d_cell_row4[0]:
                                        o_cell_row_decide=o_cell_row4[0]
                                        d_cell_row_decide=d_cell_row4[0]
                                        ws_select_bus_agency=ws_select_bus_agency2_time
                                        print("7")
                                        print(ws_select_bus_agency)
                                    else:
                                        pass
        except:
            pass

            

        print(ws_select_bus_agency)
        print(o_cell_row_decide)
        print(d_cell_row_decide)
        tram_pd0= pd.read_csv('Hakodate_tram_route.csv',encoding='UTF-8',header=None,names=['?o?X?▼','?ܓx','?o?x','???l']).fillna('-').replace(' ',0) if selected_agency=='???َs???ʋ?' else pd.read_csv('tram_stop.csv',encoding='UTF-8',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
        orig_name=ws_select_bus_agency.cell(row=o_cell_row_decide,column=1).value
        o_lat =ws_select_bus_agency.cell(row=o_cell_row_decide,column=2).value
        o_lng =ws_select_bus_agency.cell(row=o_cell_row_decide,column=3).value
        dest_name=ws_select_bus_agency.cell(row=d_cell_row_decide,column=1).value
        d_lat =ws_select_bus_agency.cell(row=d_cell_row_decide,column=2).value
        d_lng =ws_select_bus_agency.cell(row=d_cell_row_decide,column=3).value
        o={'?◯??':orig_name,'?n??':tram_pd0['???l'] if tram_pd0['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
        print(o)
        d={'?◯??':dest_name,'?n??':tram_pd0['???l'] if tram_pd0['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
        print(d)
        print("_5_5__")
        select_o_d(o,d,ws_select_bus_agency,o_cell_row_decide,d_cell_row_decide)
        

    elif selected_agency=='???ˎs?R?~???j?e?B?o?X':
        wb_select_bus_agency=openpyxl.load_workbook('setoshi_bus_record.xlsx')
        ws_select_bus_agency = wb_select_bus_agency[str(wb_select_bus_agency.sheetnames[0])]
        setoshi_lat=pd.read_csv('Setoshi_bus_stop.csv',encoding='cp932',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
        setoshi_lat=setoshi_lat[(hakodate_lat['?ܓx']!='-')|(hakodate_lat['?o?x']!='-')]
        setoshi_lat=setoshi_lat.query('0<?ܓx<90.0|?o?x!=0|?o?x!=None')
        setoshi_lat=setoshi_lat.astype({'?ܓx':float,'?o?x':float})
        setoshi_lat=setoshi_lat.to_dict(orient='records')

        if current_dep==True:
            for cell in ragne(2,ws_select_bus_agency.max_row):
                o_lat=geo_google()["location"]["lat"]
                o_lng=geo_google()["location"]["lng"]
                orig_name='???ݒn'
                o={'?◯??':orig_name,'?ܓx':o_lat,'?o?x':o_lng}
                if ws_select_bus_agency.cell(row=cell.row,column=2).value==o_lat:
                    o_cell_row=cell
                    dep_station_name=ws_select_bus_agency.cell(row=cell.row,column=1).value
                    dep_station_lat=ws_select_bus_agency.cell(row=cell.row,column=2).value
                    dep_station_lng=ws_select_bus_agency.cell(row=cell.row,column=3).value
                    if arr_stop ==True:
                        if ws_select_bus_agency.cell(row=cell.row,column=1).value==arr_stop:
                            d_cell_row=cell
                            if o_cell_row<d_cell_row:
                                dest_lat=ws_select_bus_agency.cell(row=d_cell_row,column=2).value
                                dest_lng=ws_select_bus_agency.cell(row=d_cell_row,column=3).value
                                dest_name=ws_select_bus_agency.cell(row=d_cell_row,column=1).value
                                o={'?◯??':dep_station_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dep_station_lat else None,'????':o_cell_row,'?ܓx':dep_station_lat,'?o?x':dep_station_lng}
                                d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dest_lat else None,'?ܓx':dest_lat,'?o?x':dest_lng}
                                select_o_d()
                            else:
                                ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                setoshi_lat=setoshi_lat.query('?ܓx!=@ws_select')
                                setoshi_lat=setoshi_lat.to_dict(orient='records')
                                near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                                near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                                d_temp_cell_row=cell.row
                                for i in range(d_temp_cell_row+1,ws_select_bus_agency.max_row):
                                    if ws_select_bus_agency.cell(row=i,column=2).value==near_st_orig_lat:
                                        o1_cell_row=i
                                        orig_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                        orig_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                        orig_name=ws_select_bus_agency.cell(row=o1_cell_row,column=1).value
                                        if ws_select_bus_agency.cell(row=i,column=1).value==arr_stop:
                                            d1_cell_row=i+1
                                            dest_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                            dest_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                            dest_name=ws_select_bus_agency.cell(row=d1_cell_row,column=1).value
                                            o={'?◯??':dep_station_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dep_station_lat else None,'????':o1_cell_row-35,'?ܓx':orig_lat,'?o?x':orig_lng}
                                            d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dest_lat else None,'?ܓx':dest_lat,'?o?x':dest_lng}
                                            select_o_d()
                                        else:
                                            print('pass_temp_d1')
                                    else:
                                        print('pass_temp2')
                        else:
                            print('passd_cell_d')
                    elif arr_pos==True:
                        near_st_dest_lat=nearest_stop_dest_lat(hakodate_lat_pd,d1)['?ܓx']
                        near_st_dest_lng=nearest_stop_dest_lat(hakodate_lat_pd,d1)['?o?x']
                        if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                            d_cell_row=cell.row
                            if cell.row<d_cell_row:
                                dest_lat=ws_select_bus_agency.cell(row=d_cell_row,column=2).value
                                dest_lng=ws_select_bus_agency.cell(row=d_cell_row,column=3).value
                                dest_name=ws_select_bus_agency.cell(row=d_cell_row,column=1).value
                                o={'?◯??':dep_station_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dep_station_lat else None,'????':o_cell_row,'?ܓx':dep_station_lat,'?o?x':dep_station_lng}
                                d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dest_lat else None,'?ܓx':dest_lat,'?o?x':dest_lng}
                                select_o_d()
                            else:
                                ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                setoshi_lat=hakodate_lat.query('?ܓx!=@ws_select')
                                setoshi_lat=hakodate_lat.to_dict(orient='records')
                                near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                                near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                                d_temp_cell_row=cell.row
                                for i in range(d_temp_cell_row+1,ws_select_bus_agency.max_row):
                                    if ws_select_bus_agency.cell(row=i,column=2).value==near_st_orig_lat:
                                        o1_cell_row=i
                                        if ws_select_bus_agency.cell(row=i,column=1).value==arr_stop:
                                            d1_cell_row=i+1
                                            orig_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                            orig_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                            orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                                            dest_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                            dest_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                            dest_name=ws_select_bus_agency.cell(row=d1_cell_row,column=1).value
                                            o={'?◯??':dep_station_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==orig_lat else None,'????':o1_cell_row-35,'?ܓx':orig_lat,'?o?x':orig_lng}
                                            d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==dest_lat else None,'?ܓx':dest_lat,'?o?x':dest_lng}
                                            select_o_d()
                                        else:
                                            print('pass_temp_d3')
                                    else:
                                        print('pass_temp4')
                        else:
                            print('passd_cell_d5')
                    else:
                        print('passarr_pos6')
                else:
                    pass
        elif current_dep==False:
            if dep_stop==True:
                if arr_stop==True:
                    for cell in (2,ws_select_bus_agency.max_row):
                        near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                        near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                        if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_orig_lat:
                            o_cell_row=cell.row
                            o_lat=ws_select_bus_agency.cell(row=o_cell_row,column=2).value
                            o_lng=ws_select_bus_agency.cell(row=o_cell_row,column=3).value
                            orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                            o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                            near_st_dest_lat=nearest_stop_dest_lat(hakodate_lat,d1)['?ܓx']
                            near_st_dest_lng=nearest_stop_dest_lat(hakodate_lat,d1)['?o?x']
                            if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                                d_cell_row=cell.row
                                    
                                d_lat=ws_select_bus_agency.cell(row=d_cell_row,column=2).value
                                d_lng=ws_select_bus_agency.cell(row=d_cell_row,column=3).value
                                dest_name=ws_select_bus_agency.cell(row=d_cell_row,column=1).value
                                d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                select_o_d()
                            else:
                                print('pass6.5')
                        elif ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                            d_temp_cell_row=cell.row
                            for i in range(d_temp_cell_row,ws_select_bus_agency.max_row):
                                if ws_select_bus_agency.cell(row=i,column=1).value==near_st_orig_lat:
                                    o1_cell_row=i
                                    ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                    setoshi_lat=hakodate_lat.query('?ܓx!=@ws_select')
                                    setoshi_lat=hakodate_lat.to_dict(orient='records')
                                    near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                                    near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                                    if ws_select_bus_agency.cell(row=o1_cell_row,column=2).value==near_st_orig_lat:
                                        o_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                        o_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                        orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                                        o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                        near_st_dest_lat=nearest_stop_o_lat(hakodate_lat,d1)['?ܓx']
                                        near_st_dest_lng=nearest_stop_o_lat(hakodate_lat,d1)['?o?x']
                                        if ws_select_bus_agency.cell(row=i,column=2).value==near_st_dest_lat:
                                            d1_cell_row=i
                                            d_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                            d_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                            dest_name=ws_select_bus_agency.cell(row=d1_cell_row-35,column=1).value
                                            
                                            d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                            select_o_d()
                                        else:
                                            pass
                                    else:
                                        print('pass_d7')
                                else:
                                    print('passarr_posarr_pos8')
                        else:
                            print('passarr_posarr_pos9')
                            ###
                elif arr_pos==True:
                    for col in ws_select_bus_agency.iter_cols(min_row=2,max_row=ws_select_bus_agency.max_row,min_col=1,max_col=ws_select_bus_agency.max_column):
                        for cell in col:
                            near_st_dest_lat=nearest_stop_dest_lat(hakodate_lat,d1)['?ܓx']
                            near_st_dest_lng=nearest_stop_dest_lat(hakodate_lat,d1)['?o?x']
                            near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                            near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                            if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_orig_lat:
                                o_lat=near_st_orig_lat
                                o_lng=near_st_orig_lng
                                orig_name=nearest_stop_o_lng(hakodate_o_lat,d1)['?o?X?▼']
                                o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                        
                                
                                print(o)
                                if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_orig_lat and ws_select_bus_agency.cell(row=cell.row,column=3).value== near_st_dest_lng:
                                    d_lat=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                    d_lng=ws_select_bus_agency.cell(row=cell.row,column=3).value
                                    dest_name=ws_select_bus_agency.cell(row=cell.row,column=1).value
                                    d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                    select_o_d()
                                    print(d)
                                else:
                                    print('pass_d10')
                            elif ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                                ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                setoshi_lat=hakodate_lat.query('?ܓx!=@ws_select')
                                setoshi_lat=hakodate_lat.to_dict(orient='records')
                                near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                                near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                                d_temp_cell_row=cell.row
                                for i in range(d_temp_cell_row,ws_select_bus_agency.max_row):
                                    if ws_select_bus_agency.cell(row=i,column=1).value==near_st_orig_lat:
                                        o1_cell_row=i
                                        o_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                        o_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                        orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                                        o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                        if ws_select_bus_agency.cell(row=i,column=2).value==near_st_dest_lat:
                                            d1_cell_row=i
                                            d_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                            d_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                            dest_name=ws_select_bus_agency.cell(row=d1_cell_row-35,column=1).value
                                            
                                            d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                            select_o_d()
                                        else:
                                            pass
                                    else:
                                        print('pass_d11')
                            else:
                                print('passarr_posarr_pos112')
                else:
                    print('passarr_posarr_posarr_pos13')
            elif dep_pos== True:
                if arr_stop==True:
                    for col in ws_select_bus_agency.iter_cols(min_row=2,max_row=ws_select_bus_agency.max_row,min_col=1,max_col=ws_select_bus_agency.max_column):
                        for cell in col:
                            near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                            near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                            print(near_st_orig_lat)
                            time.sleep(3)
                            if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_orig_lat:
                                o_cell_row=cell.row
                                o_lat=ws_select_bus_agency.cell(row=o_cell_row,column=2).value
                                o_lng=ws_select_bus_agency.cell(row=o_cell_row,column=3).value
                                orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                                o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                print(o)
                                time.sleep(3)
                                if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                                    d_cell_row=cell.row
                                    d_lat=ws_select_bus_agency.cell(row=d_cell_row,column=2).value
                                    d_lng=ws_select_bus_agency.cell(row=d_cell_row,column=3).value
                                    dest_name=ws_select_bus_agency.cell(row=d_cell_row,column=1).value
                                    d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                    print(d)
                                    select_o_d()
                                else:
                                    print('pass')
                            else:
                                d_temp_cell_row=cell.row
                                for i in range(d_temp_cell_row,ws_select_bus_agency.max_row):
                                    if ws_select_bus_agency.cell(row=i,column=2).value==near_st_orig_lat:
                                        o1_cell_row=i
                                        ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                                        setoshi_lat=hakodate_lat.query('?ܓx!=@ws_select')
                                        setoshi_lat=hakodate_lat.to_dict(orient='records')
                                        near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                                        near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                                        if ws_select_bus_agency.cell(row=o1_cell_row,column=2).value==near_st_orig_lat:
                                            o_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                            o_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                            orig_name=ws_select_bus_agency.cell(row=o_cell_row,column=1).value
                                            o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                            near_st_dest_lat=nearest_stop_dest_lat(hakodate_lat,d1)['?ܓx']
                                            near_st_dest_lng=nearest_stop_dest_lat(hakodate_lat,d1)['?o?x']
                                            print(o)
                                            if ws_select_bus_agency.cell(row=i,column=2).value==near_st_dest_lat:
                                                d1_cell_row=i
                                                d_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                                d_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                                dest_name=ws_select_bus_agency.cell(row=d1_cell_row-35,column=1).value
                                            
                                                d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                                print(d)
                                                select_o_d()
                                            else:
                                                pass
                                        else:
                                            print('pass_d14')
                                    else:
                                        print('passarr_posarr_pos15')
                            ###
                elif arr_pos==True:
                    for cell in range(2, ws_select_bus_agency.max_row):
                        near_st_dest_lat=nearest_stop_dest_lat(hakodate_lat,d1)['?ܓx']
                        near_st_dest_lng=nearest_stop_dest_lat(hakodate_lat,d1)['?o?x']
                        near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                        near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                        print(cell.row)
                        if ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_orig_lat:
                            o_lat=near_st_orig_lat
                            o_lng=near_st_orig_lng
                            orig_name=nearest_stop_orig_lat(hakodate_lat,o1)['?o?X?▼']
                            o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}

                                
                            print(o)
                            for i in range(cell.row,ws_select_bus_agency.max_row):
                                if ws_select_bus_agency.cell(row=i,column=2).value==near_st_dest_lat and ws_select_bus_agency.cell(row=i,column=3).value== near_st_dest_lng:
                                    d_lat=ws_select_bus_agency.cell(row=i,column=2).value
                                    d_lng=ws_select_bus_agency.cell(row=i,column=3).value
                                    dest_name=nearest_stop_dest_lat(hakodate_lat,d1)['?o?X?▼']
                                    d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                    print(d)
                                    select_o_d()
                                else:
                                    print('pass_d17')
                                break
                        elif ws_select_bus_agency.cell(row=cell.row,column=2).value==near_st_dest_lat:
                            ws_select=ws_select_bus_agency.cell(row=cell.row,column=2).value
                            setoshi_lat=pd.read_csv('Setoshi_bus_stop.csv',encoding='cp932',header=None,names=['???Ǝ?','?o?X?▼','?o?X?▼(?J?i)','?ܓx','?o?x','???l']).fillna('-').replace(' ',0)
                            setoshi_lat=hakodate_lat[(hakodate_lat['?ܓx']!='-')|(hakodate_lat['?o?x']!='-')]
                            setoshi_lat=hakodate_lat.query('0<?ܓx<90.0|?o?x!=0|?o?x!=None')
                            setoshi_lat=hakodate_lat.astype({'?ܓx':float,'?o?x':float})
                            setoshi_lat=hakodate_lat.query('?ܓx!=@ws_select')
                            setoshi_lat=hakodate_lat.to_dict(orient='records')
                            near_st_orig_lat=nearest_stop_orig_lat(hakodate_lat,o1)['?ܓx']
                            near_st_orig_lng=nearest_stop_orig_lat(hakodate_lat,o1)['?o?x']
                            d_temp_cell_row=cell.row
                            for i in range(d_temp_cell_row,ws_select_bus_agency.max_row):
                                if ws_select_bus_agency.cell(row=i,column=1).value==near_st_orig_lat:
                                    o1_cell_row=i
                                    o_lat=ws_select_bus_agency.cell(row=o1_cell_row,column=2).value
                                    o_lng=ws_select_bus_agency.cell(row=o1_cell_row,column=3).value
                                    orig_name=nearest_stop_orig_lat(hakodate_lat,o1)['?o?X?▼']
                                    o={'?◯??':orig_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==o_lat else None,'?ܓx':o_lat,'?o?x':o_lng}
                                    print(o)
                                    if ws_select_bus_agency.cell(row=i,column=2).value==near_st_dest_lat:
                                        d1_cell_row=i
                                        d_lat=ws_select_bus_agency.cell(row=d1_cell_row,column=2).value
                                        d_lng=ws_select_bus_agency.cell(row=d1_cell_row,column=3).value
                                        dest_name=nearest_stop_dest_lat(hakodate_lat,d1)['?o?X?▼']
                                            
                                        d={'?◯??':dest_name,'?n??':tram_pd['???l'] if tram_pd['?ܓx'].any()==d_lat else None,'?ܓx':d_lat,'?o?x':d_lng}
                                        print(d)
                                        select_o_d()
                                    else:
                                        pass
                                else:
                                    print('pass_d18')
                        else:
                            print('pass_d19')
            else:
                print('passarr_posarr_posarr_pos20')
        else:
            pass
    else:
        pass
